
import openpyxl
import time
import sys
import pathlib
import requests

from selenium import webdriver

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options


message_request = {
    "subject":"hi",
    "message":"hi",
    "attachments":"{}",
    "recipient_loginname":"Dasosmijewels",
    "recipient_id":null,
    "captcha_code":"g-recaptcha-convos-compose-64760dac1e61a",
    "g-recaptcha-response":"",
    "api_context":"buyer_conversations"
}


class MsgSender:

    def __init__(self, filepath, username, password, msg_template,
            start_num, num_msg):

        self.username = username
        self.password = password
        # self.url_col = url_col
        self.msg_template = msg_template
        self.start_num = start_num
        self.num_msg = num_msg

        wb_obj = openpyxl.load_workbook(filepath)
        self.sheet_obj = wb_obj.active

        chrome_options = Options()
        if sys.platform == "linux":
            chrome_options.add_argument("user-data-dir=selenium")
        else:
            script_dir = pathlib.Path().absolute()
            chrome_options.add_argument(f"user-data-dir={script_dir}\\selenium")

        self.driver = webdriver.Chrome(options=chrome_options)

    def get_sellers(self):
        i = self.start_num + 1
        max_i = i + self.num_msg
        urls = []
        names = []
        messages = []
        url_col = 2

        for k in range(2, 100):
            col_val = self.sheet_obj.cell(row=2, column=k).value
            print("COL VAL", col_val)
            if col_val and "etsy.com" in col_val:
                if (col_val.startswith("http") or col_val.startwith("www")
                        or col_val.startswith("etsy")):
                    url_col = k
                    break
        print("URL COL", url_col)
                
        while self.sheet_obj.cell(row=i, column=url_col).value:
            url = self.sheet_obj.cell(row=i, column=url_col).value
            params = []
            j = 1
            while self.sheet_obj.cell(row=i, column=url_col + j).value:
                params.append(self.sheet_obj.cell(
                    row=i, column=url_col + j).value)
                j += 1
                
            # message = self.sheet_obj.cell(row=i, column=msg_col).value
            print("Params", params)
            params += ['' * 1000]
            message = self.msg_template.format(*params)
            urls.append(url)
            names.append(urls.rsplit("/", 1)[-1])
            messages.append(message)
            i += 1
            if i >= max_i:
                break
        return urls, names, messages
    
    def safe_evaluate(self, script, tries=1):
        while tries > 0:
            try:
                return self.driver.execute_script(script)
            except Exception as err:
                print("Evaluation Error:", err)
                if "Alert Text" in str(err):
                    tries += 1
            tries -= 1

    def is_captcha_solved(self):
        return self.safe_evaluate('return (typeof grecaptcha !== "undefined") && grecaptcha && grecaptcha.enterprise?.getResponse().length > 0', tries=3)


    def set_textarea(self, selector, value):
        value = value.replace('"', '\\"').replace("'", "\\'")
        self.safe_evaluate(
            "var valSetter=Object.getOwnPropertyDescriptor("
            "window.HTMLTextAreaElement.prototype, 'value').set;"
            f"var input = document.querySelector('{selector}');"
            f"valSetter.call(input,'{value}');"
            "var ev = new Event('input',{bubbles:true});"
            "input.dispatchEvent(ev);", 3)

    def is_browser_open(self):
        try:
            self.driver.current_url
            return EC.alert_is_present() 
        except:
            return False

    def start_sending(self):
        seller_pages, seller_names, messages = self.get_sellers()

        driver = self.driver
        driver.get("https://www.etsy.com")

        wait = WebDriverWait(driver, 180)

        if not self.safe_evaluate('return document.querySelector(".welcome-message-text")', 3):
            
            print("NO Login History")
            driver.execute_script('document.querySelector("button.select-signin").click()')
            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#join_neu_email_field')))

            driver.execute_script(f'document.querySelector("#join_neu_email_field").value = "{self.username}"')
            driver.execute_script(f'document.querySelector("#join_neu_password_field").value = "{self.password}"')
            if self.password.strip() and self.username.strip():
                driver.execute_script('document.querySelector("button[name=\'submit_attempt\']").click()')

            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.welcome-message-text')))
        print("Logged IN")

        driver.get("

        
        driver.close()
        return 0

