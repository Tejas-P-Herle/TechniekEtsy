#!/usr/bin/env python3

import openpyxl
from selenium import webdriver

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()

email = "yifefa3387@introace.com"
username = "InternshalaTask"
password = "Task@Internshala"

path = "sellers.xlsx"
message = "FORMATTABLE_MESSAGE_TEMPALTE"
url_column = 3
msg_column = url_column - 1


def get_sellers(sheet_obj):
    i = 2
    urls = []
    messages = []
    while sheet_obj.cell(row=i, column=url_column).value:
        url = sheet_obj.cell(row=i, column=url_column).value
        message = sheet_obj.cell(row=i, column=msg_column).value
        urls.append(url)
        messages.append(message)
        i += 1
    return urls, messages


def main():
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    seller_pages, messages = get_sellers(sheet_obj)

    driver.get("https://www.etsy.com")

    wait = WebDriverWait(driver, 180)

    driver.execute_script('document.querySelector("button.select-signin").click()')
    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#join_neu_email_field')))

    driver.execute_script(f'document.querySelector("#join_neu_email_field").value = "{email}"')
    driver.execute_script(f'document.querySelector("#join_neu_password_field").value = "{password}"')
    driver.execute_script('document.querySelector("button[name=\'submit_attempt\']").click()')

    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.welcome-message-text')))
    print("Logged IN")
    for i, page in enumerate(seller_pages):
        driver.get(page)
        driver.execute_script('document.querySelector("#desktop_shop_owners_parent").querySelector("a.wt-btn.wt-btn--outline.wt-width-full.contact-action.convo-overlay-trigger.inline-overlay-trigger").click()')
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "textarea[placeholder='Write a message']")))
        mod_message = messages[i]
        driver.execute_script(f'document.querySelector("textarea[placeholder=\'Write a message\']").value = "{mod_message}"')
        print(f"Send Message: '{mod_message}' to {page}")
        input("Click ENTER to goto next page")

    driver.close()

     
if __name__ == "__main__":
    main()

