#!/usr/bin/env python3

import openpyxl
import time
import os

from tkinter import *
from os import path
from tkinter.filedialog import askopenfilename

import requests
import grequests
import json
import sys
import subprocess
import signal

from subprocess import PIPE

from lxml.html import fromstring
from selenium import webdriver

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options


returncode = subprocess.call(["sudo", "echo", "HI"])


cookies = {}
csrf_token = ""
page_guid = ""

input_ok = False

filepath = "sellers_names.xlsx"
filepath = input("Filepath(sellers_names.xlsx): ")
if not filepath.strip():
    filepath = "sellers_names.xlsx"

message_file = input("Message File(Default Message - 'Hi'): ")
if not message_file.strip():
    message = "HI"
else:
    with open(message_file) as file:
        message = file.read()

message = message.replace('"', '\\"').replace("'", "\\'")

input_ok = True
print("Message", message)
num_msg = 20
# filepath = ""
# message = "HI"


batch_size = 30
TIMEOUT=20

sleep_time = 30
row = 0

contacted_sellers = set()

ovpn_files = []
messages_left = 0
max_session_msg = 3


def set_cookies(driver):
    global cookies

    all_cookies = driver.get_cookies()
    cookies = {}
    for cookie in all_cookies:
        cookies[cookie['name']] = cookie['value']
    print("COOKIES", cookies)


def set_codes(driver):
    global csrf_token, page_guid

    csrf_token = driver.execute_script(
        "return document.querySelector('meta[name=\"csrf_nonce\"]').content")
    page_guid = driver.execute_script("guid_start = document.body.innerHTML.indexOf('page_guid\":') + 12; return document.body.innerHTML.slice(guid_start, document.body.innerHTML.indexOf('\"', guid_start+1))")


def get_captcha(driver):
    try:
        return driver.execute_async_script("""
            async function getParams() {
              let resp = await fetch("https://www.etsy.com/api/v3/ajax/member/conversations/recaptcha-data");
              let respJson = await resp.json();

              let captcha_code = respJson["id"];
                
              gresp = grecaptcha.enterprise.getResponse();
              
              return [captcha_code, gresp]
            }
            params = getParams()
            params.then(resp => arguments[0](resp))
        """)
    except:
        return "", ""

    
def sendMsgCore(user, message, captcha_code, recaptchaResp, timeout=TIMEOUT):

    reqJson = {
        "subject": "hi",
        "message": message,
        "attachments": "{}",
        "recipient_loginname": user,
        "recipient_id": None,
        "captcha_code": captcha_code,
        "g-recaptcha-response": recaptchaResp,
        "api_context": "buyer_conversations"
    }

    resp = requests.post(
        "https://www.etsy.com/api/v3/ajax/member/conversations",
        json=reqJson, headers={
          "x-csrf-token": csrf_token,
          "x-detected-locale": "INR|en-GB|IN",
          "page_guid": page_guid
        }, cookies=cookies, timeout=timeout
    )

    return resp


def sendMsgProxyPool(user, message, captcha_code, recaptchaResp, proxyBatch, timeout=TIMEOUT):

    reqJson = {
        "subject": "hi",
        "message": message,
        "attachments": "{}",
        "recipient_loginname": user,
        "recipient_id": None,
        "captcha_code": captcha_code,
        "g-recaptcha-response": recaptchaResp,
        "api_context": "buyer_conversations"
    }
    
    rs = []
    for proxy in proxyBatch:
        proxies = {"http": proxy, "https": proxy}
        rs.append(grequests.post(
            "https://www.etsy.com/api/v3/ajax/member/conversations",
            json=reqJson, headers={
              "x-csrf-token": csrf_token,
              "x-detected-locale": "INR|en-GB|IN",
              "page_guid": page_guid
            }, cookies=cookies, proxies=proxies, timeout=timeout
        ))

    return grequests.map(rs)

def reset_captcha(driver, captcha_code):
    
    return driver.execute_script(
        f"document.querySelector('#{captcha_code}'); grecaptcha.enterprise.reset();")


def sendMsg(user, message, proxyPool=False, timeout=TIMEOUT):
    captcha_code, gresp = get_captcha(driver)

    if proxyPool is False:
        msg_resp = sendMsgCore(user, message, captcha_code, gresp, timeout=timeout)
        # reset_captcha(driver, captcha_code)
        return msg_resp, captcha_code

    src = 'https://api64.ipify.org'
    response = requests.get('https://free-proxy-list.net/')
    parser = fromstring(response.text)
    proxies = []

    for i in parser.xpath('//tbody/tr'):
        if i.xpath('.//td[7][contains(text(),"yes")]'):
            proxy = ":".join([i.xpath('.//td[1]/text()')[0],
                              i.xpath('.//td[2]/text()')[0]])
            proxies.append(proxy)

    print("Got Proxies")
    batches = [proxies[s:s+batch_size] for s in range(0, len(proxies), batch_size)]
    for ix, batch in enumerate(batches):
        print("Testing Batch", ix)
        resps = sendMsgProxyPool(user, message, captcha_code, gresp, batch, timeout)
        print("Batch Done", ix, resps)
        # print("RESPS", resps)
        for resp in resps:
            if resp and resp.status_code == 201:
                # reset_captcha(driver, captcha_code)
                return resp, captcha_code
    return None, ""


def get_working_proxies(req_no=5):
    src = 'https://api64.ipify.org'
    response = requests.get('https://free-proxy-list.net/')
    parser = fromstring(response.text)
    proxies = []

    for i in parser.xpath('//tbody/tr'):
        if i.xpath('.//td[7][contains(text(),"yes")]'):
            proxy = ":".join([i.xpath('.//td[1]/text()')[0],
                              i.xpath('.//td[2]/text()')[0]])
            proxies.append(proxy)

    working_proxies = []
    batches = [proxies[s:s+batch_size] for s in range(0, len(proxies), batch_size)]
    for batch in batches:
        print("Testing Batch")
        rs = (grequests.get(src, proxies={'http': proxy, 'https': proxy}, timeout=20) for u in batch)
        for i, resp in enumerate(grequests.map(rs)):
            if resp and resp.ok:
                working_proxies.append(batch[i])
        print("Batch Complete, Working:", len(working_proxies))
        if len(working_proxies) >= req_no:
            break
    print("Working Proxies:", working_proxies)

    return working_proxies


def is_captcha_solved(driver):
    try:
        return driver.execute_script(
            'return (typeof grecaptcha !== "undefined") && grecaptcha && grecaptcha.enterprise?.getResponse().length > 0')
    except:
        return False
    

def is_browser_open(driver):
   try:
        driver.window_handles
        return True
   except:
        return False


def get_sellers(sheet_obj):
    sellers = []
    i = 1 + 1
    max_i = i + num_msg
    seller_col = 1
    print("Seller COL", seller_col)

    for k in range(1, 100):
        col_val = sheet_obj.cell(row=1, column=k).value
        if "Seller" in col_val or "Name" in col_val:
            seller_col = k
            break
    
    for k in range(i, max_i):
        if sheet_obj.cell(row=k, column=seller_col).value:
            seller = sheet_obj.cell(row=k, column=seller_col).value
            sellers.append(seller)
    return sellers


def add_label(root, label_text, column=0, same_row=False, font=None, **kwargs):
    global row
    
    label = Label(root, text=label_text, font=font)
    label.grid(row=row, column=column, **kwargs)
    row += not same_row

    return label

def add_entry(root, width=40, column=0, padx=5, same_row=False,
              num_only=False, default=None, show=None, **kwargs):
    global row

    vcmd = None
    if num_only:
        vcmd = (root.register(validate),
                        '%d', '%i', '%P', '%s', '%S', '%v', '%V', '%W')
    entry = Entry(root, width=width, validate = 'key', show=show,
        validatecommand=vcmd)
    if default is not None:
        entry.insert(END, str(default))

    entry.grid(row=row, column=column, padx=padx, **kwargs)
    row += not same_row

    return entry

def add_button(root, btn_text, command, column=1,
               same_row=False, **kwargs):
    global row
    button = Button(root, text=btn_text, command=command)
    button.grid(row=row-1, column=column, sticky="nesw", **kwargs)
    row += not same_row

    return button

def add_text(root, height=5, width=52, column=0, columnspan=2,
             padx=5, pady=5, same_row=False, **kwargs):
    global row
    text = Text(root, height=height, width=width)
    text.grid(row=row, column=column, columnspan=columnspan,
              padx=padx, pady=pady)
    row += not same_row

    return text

def validate(action, index, value_if_allowed,
                   prior_value, text, validation_type, trigger_type, widget_name):
    if value_if_allowed:
        try:
            int(value_if_allowed)
            return True
        except ValueError:
            return False
    else:
        return True


def setExcelFile(excelFileInput):
    excelFilePath = askopenfilename(filetypes=[("Excel files", ".xlsx .xls")])
    excelFileInput.delete(0,END)
    excelFileInput.insert(0,excelFilePath)


def start_sending(root, input_fields):
    global status_label, input_ok, filepath, message, num_msg

    def getValue(widget):
        if isinstance(widget, Text):
            text = widget.get("1.0", END)
            text = text.strip().replace("\n", "\\n").replace("\t", "\\t")
            return text
        else:
            val = widget.get()
            if val.isdigit():
                return int(val)
            elif val.isdecimal():
                return float(val)
            return val
    status_label.config(text="Status: Sending Messages")
    input_values = {field_name: getValue(widget)
        for field_name, widget in input_fields.items()}
    print("input_values", input_values)
    
    if not input_values["num_msg"]:
        input_values["num_msg"] = 100

    if not input_values["filepath"] or not path.exists(input_values["filepath"]):
        status_label.config(text="Status: Invalid filepath")
        print("Invalid FilePath")
        return

    if not input_values["msg_template"]:
        status_label.config(text="Status: Empty Message")
        print("Empty Message")
        return

    filepath = input_values["filepath"]
    message = input_values["msg_template"].replace('"', '\\"').replace("'", "\\'")
    num_msg = input_values["num_msg"]
    
    input_ok = True
    root.destroy()


def load_ovpns():
    global ovpn_files
    for file in os.listdir("vpns"):
        if file.endswith(".ovpn"):
            ovpn_files.append(file)
    ovpn_files = sorted(ovpn_files)
    print("ovpn files", ovpn_files)


def kill_all_vpns():
    p = subprocess.Popen(['sudo', 'killall', 'openvpn'],
            stdout=PIPE, stdin=PIPE, stderr=PIPE)

def get_ip():
    try:
        return requests.get("https://api.ipify.org/", timeout=5).text
    except:
        return None
    

def rotate_vpn():
    print("Rotate VPN")
    while True:
        print("Load OVPN")
        load_ovpns()
        print("Done Loading")
        prev_ip = get_ip()
        for vpn in ovpn_files:
            print("Connecting to VPN")
            p = subprocess.Popen(['sudo', 'openvpn', "vpns/" + vpn],
                stdin=PIPE, stderr=PIPE) #, stdout=PIPE)
            resp = None
            count = 20
            retry = False
            while not resp or resp == prev_ip:
                time.sleep(1)
                resp = get_ip()
                count -= 1
                if count == 0:
                    retry = True
                    break
                print("RESP", resp)

            if retry:
                kill_all_vpns()
                continue
            
            print("Connected, ip", resp)
            yield

            # prev_ip = get_ip()
            prev_ip = None
            kill_all_vpns()
            print("VPN Closed")

def build_app(root):
    global row

    root.title("Techneik Etsy Messenger")
    
    heading_label = add_label(root, "Etsy Auto-Messenger", columnspan=2, font=24)
    status_label = add_label(root, "Status: Loaded", columnspan=2)

    col1_ewidth = 16
    sel_label = add_label(root, "Select Excel File:")
    filepath_entry = add_entry(root)
    load_button = add_button(root, "Select", lambda: setExcelFile(filepath_entry))
    msg_label = add_label(root, "Enter Message:", columnspan=2)
    
    message_box = add_text(root)

    start_label = add_label(root, "Starting Seller Offset", same_row=True)
    num_label = add_label(root, "Number of Messages", same_row=True)
    num_msg_entry = add_entry(root, width=col1_ewidth, column=1,
        num_only=True, default=3)

    input_fields = {'filepath': filepath_entry,
        'msg_template': message_box, 'num_msg': num_msg_entry}

    row += 1
    send_msg_btn = add_button(root, "Send Messages",
        lambda: start_sending(root, input_fields),
        column=0, columnspan=2, padx=5, pady=5)


def open_browser(chrome_options):
    driver = webdriver.Chrome(options=chrome_options)

    driver.get("https://www.etsy.com/messages/sent")
    wait = WebDriverWait(driver, 1800)
    print("Loaded, Searching for button")
    waitS = WebDriverWait(driver, 7)
    try:
        waitS.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "[data-test-id=convo-compose-button]")))
    except:
        return driver, False

    driver.execute_script("document.querySelector('[data-test-id=convo-compose-button]').click()")

    set_cookies(driver)
    set_codes(driver)

    return driver, True
    

def main():
    global driver, status_label, row, messages_left

    # root = Tk()
    # build_app(root)
    # root.mainloop()

    if not input_ok:
        print("Invalid Input")
        return

    wb_obj = openpyxl.load_workbook(filepath)
    sheet_obj = wb_obj.active

    sellers = get_sellers(sheet_obj)

    print("Sellers", sellers)
    
    chrome_options = Options()
    if "linux" in sys.platform:
        chrome_options.add_argument("user-data-dir=selenium")
    elif "win32" in sys.platform:
        script_dir = pathlib.Path().absolute()
        chrome_options.add_argument(f"user-data-dir={script_dir}\\selenium")
    else:
        script_dir = pathlib.Path().absolute()
        chrome_options.add_argument(f"user-data-dir={script_dir}/selenium")

    try:
        with open("contacted_sellers.json") as file:
            contacted_sellers = json.load(file)
    except:
        contacted_sellers = []
    
    driver = None
    vpn_rotator = rotate_vpn()
    try:
        for seller in sellers:
            if seller in contacted_sellers:
                print("Already Contacted", seller)
                continue
            if messages_left < 1:
                print("Loop Next")
                if driver:
                    driver.quit()
                # next(vpn_rotator)
                print("Open Browser")
                stat_ok = False
                while stat_ok is False:
                    next(vpn_rotator)
                    driver, stat_ok = open_browser(chrome_options)
                    if not stat_ok:
                        driver.quit()
                messages_left = max_session_msg

            while not is_captcha_solved(driver):
                if not is_browser_open(driver):
                    print("Browser Closed")
                    return
                time.sleep(0.1)
            
            print("Sending Message")
            resp, captcha_code = sendMsg(seller, message)
            # resp = None
            # captcha_code, gresp = get_captcha(driver)
            if resp and resp.status_code == 201:
                print("Message sent successfully to", seller)
                contacted_sellers.append(seller)
                messages_left -= 1
            else:
                print("Resp", resp)
                if resp is not None:
                    if resp.status_code == 200:
                        print("Rate Limit HIT")
                        messages_left = 1
                    print("Content", resp.content)
                print("Failed To Send Message To", seller)

                messages_left -= 1

            # time.sleep(sleep_time)
            if captcha_code and captcha_code.strip():
                reset_captcha(driver, captcha_code)
    finally:
        with open("contacted_sellers.json", "w") as file:
            json.dump(contacted_sellers, file, indent=2)

        if driver:
            driver.quit()
        kill_all_vpns()
        subprocess.call(['stty', 'echo'])


if __name__ == "__main__":
    main()

